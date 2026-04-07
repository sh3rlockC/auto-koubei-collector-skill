#!/usr/bin/env python3
"""导出汽车之家车型口碑为按来源链接对齐的 Excel。

功能：
- 抓取指定车型的最满意 / 最不满意口碑页
- 按来源链接严格对齐同一条口碑
- 统一判定为车主购车口碑 / 试驾探店口碑
- 导出为 2 个 sheet：购车口碑 / 试驾探店口碑
- 每行包含同一链接的“最满意 / 最不满意”两列

依赖：
- agent-browser 可用
- openpyxl 已安装
"""

import argparse
import base64
import hashlib
import hmac
import json
import re
import subprocess
import sys
import time
from html import unescape
from pathlib import Path
from datetime import datetime, timezone
from urllib import request as urllib_request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

try:
    from tqdm.auto import tqdm
except Exception:  # pragma: no cover - optional dependency fallback
    tqdm = None

HEADERS = [
    '数据类型', '用户名', '发表日期', '口碑标题', '综合口碑', '车型', '行驶里程', '电耗',
    '裸车购买价', '参考价格', '购买时间', '探店时间', '购买地点', '探店地点',
    '最满意', '最不满意', '来源链接', '最满意抓取页码', '最不满意抓取页码'
]

STAGE_WEIGHTS = {
    "detect_pages": 30,
    "collect_sat": 400,
    "collect_unsat": 400,
    "write_excel": 70,
    "write_validation": 50,
}


def progress_iter(iterable, *, total=None, desc="", enabled=False):
    if not enabled or tqdm is None:
        return iterable
    return tqdm(iterable, total=total, desc=desc, dynamic_ncols=True, leave=False, file=sys.stderr)


def utc_now_iso():
    return datetime.now(timezone.utc).astimezone().replace(microsecond=0).isoformat()


def write_json_atomic(path, payload):
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = target.with_suffix(target.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(target)


def post_json(url, payload, timeout=5):
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib_request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib_request.urlopen(req, timeout=timeout) as resp:
        resp.read()


def make_feishu_signature(secret, timestamp):
    string_to_sign = f"{timestamp}\n{secret}"
    digest = hmac.new(string_to_sign.encode("utf-8"), digestmod=hashlib.sha256).digest()
    return base64.b64encode(digest).decode("utf-8")


def build_feishu_progress_text(payload):
    percent = int(payload.get("percent", 0))
    stage = payload.get("stage") or ""
    stage_progress = payload.get("stage_progress") or {}
    current = int(stage_progress.get("current", payload.get("current", 0)))
    total = int(stage_progress.get("total", payload.get("total", 1)) or 1)
    message = payload.get("message") or ""
    model_name = payload.get("model_name") or ""
    mode = payload.get("mode") or ""
    mode_label = {
        "autohome_collect": "汽车之家采集",
    }.get(mode, mode or "任务")
    lines = [f"口碑采集 {percent}%｜{model_name}", f"{mode_label} · {stage} {current}/{total}"]
    if message and message not in {stage, f"{stage} {current}/{total}"}:
        lines.append(f"说明：{message}")
    return "\n".join(lines)


def build_feishu_payload(payload, *, secret=None):
    timestamp = int(time.time())
    feishu_payload = {
        "timestamp": str(timestamp),
        "msg_type": "text",
        "content": {
            "text": build_feishu_progress_text(payload),
        },
    }
    if secret:
        feishu_payload["sign"] = make_feishu_signature(secret, timestamp)
    return feishu_payload


class ProgressReporter:
    def __init__(
        self,
        *,
        label,
        model_name,
        mode,
        progress_file=None,
        progress_webhook=None,
        feishu_webhook=None,
        feishu_secret=None,
        total_units=1000,
    ):
        self.label = label
        self.model_name = model_name
        self.mode = mode
        self.progress_file = Path(progress_file).resolve() if progress_file else None
        self.progress_webhook = progress_webhook
        self.feishu_webhook = feishu_webhook
        self.feishu_secret = feishu_secret
        self.total_units = total_units
        self.current_units = 0
        self.stage_base_units = 0
        self.stage_weight_units = 0
        self.stage_total = 1
        self.stage_current = 0
        self.stage_key = "init"
        self.stage_label = "准备"
        self.last_webhook_percent = None
        self.last_webhook_stage = None
        self.disabled_file = False
        self.disabled_webhook = False
        self.disabled_feishu = False

    def _build_payload(self, percent=None, message=""):
        percent = int(percent if percent is not None else round(self.current_units * 100 / self.total_units))
        stage_total = max(int(self.stage_total), 1)
        stage_current = min(int(self.stage_current), stage_total)
        return {
            "label": self.label,
            "mode": self.mode,
            "model_name": self.model_name,
            "stage": self.stage_label,
            "stage_key": self.stage_key,
            "current": int(self.current_units),
            "total": int(self.total_units),
            "percent": percent,
            "message": message,
            "updated_at": utc_now_iso(),
            "overall": {
                "current": int(self.current_units),
                "total": int(self.total_units),
                "percent": percent,
            },
            "stage_progress": {
                "current": stage_current,
                "total": stage_total,
                "percent": int(round(stage_current * 100 / stage_total)),
            },
            "progress_file": str(self.progress_file) if self.progress_file else None,
            "webhook": self.progress_webhook,
        }

    def _emit(self, payload):
        if self.progress_file and not self.disabled_file:
            try:
                write_json_atomic(self.progress_file, payload)
            except Exception as exc:
                self.disabled_file = True
                print(f"WARNING: progress file write failed: {exc}", file=sys.stderr)

        percent = payload["percent"]
        should_emit_webhook = (
            percent != self.last_webhook_percent
            or payload["stage"] != self.last_webhook_stage
            or payload["message"] in {"导出完成", "输入校验失败", "校验失败"}
        )
        if not should_emit_webhook:
            return

        self.last_webhook_percent = percent
        self.last_webhook_stage = payload["stage"]

        if self.progress_webhook and not self.disabled_webhook:
            try:
                post_json(self.progress_webhook, payload)
            except Exception as exc:
                self.disabled_webhook = True
                print(f"WARNING: progress webhook failed: {exc}", file=sys.stderr)
        if self.feishu_webhook and not self.disabled_feishu:
            try:
                post_json(self.feishu_webhook, build_feishu_payload(payload, secret=self.feishu_secret))
            except Exception as exc:
                self.disabled_feishu = True
                print(f"WARNING: feishu webhook failed: {exc}", file=sys.stderr)

    def start_stage(self, stage_key, stage_label, *, weight, total=1, message=""):
        self.stage_key = stage_key
        self.stage_label = stage_label
        self.stage_base_units = self.current_units
        self.stage_weight_units = max(int(weight), 0)
        self.stage_total = max(int(total), 1)
        self.stage_current = 0
        self._emit(self._build_payload(message=message or stage_label))

    def advance(self, current, *, message=""):
        self.stage_current = max(0, min(int(current), self.stage_total))
        fraction = self.stage_current / self.stage_total if self.stage_total else 1.0
        self.current_units = min(self.total_units, int(round(self.stage_base_units + self.stage_weight_units * fraction)))
        self._emit(self._build_payload(message=message or self.stage_label))

    def finish_stage(self, message=""):
        self.stage_current = self.stage_total
        self.current_units = min(self.total_units, int(round(self.stage_base_units + self.stage_weight_units)))
        self._emit(self._build_payload(message=message or f"{self.stage_label}完成"))

    def finalize(self, stage_label="完成", message="导出完成"):
        self.stage_label = stage_label
        self.stage_key = stage_label
        self.stage_current = self.stage_total
        self.current_units = self.total_units
        payload = self._build_payload(percent=100, message=message)
        payload["stage_progress"]["current"] = payload["stage_progress"]["total"]
        payload["stage_progress"]["percent"] = 100
        self._emit(payload)


def run(cmd: str, cwd: Path, timeout: int = 120):
    return subprocess.run(cmd, shell=True, capture_output=True, text=True, cwd=str(cwd), timeout=timeout)


def url_for(series_id: int, dimensionid: int, page: int) -> str:
    if page == 1:
        return f'https://k.autohome.com.cn/{series_id}?dimensionid={dimensionid}&order=0&yearid=0#listcontainer'
    return f'https://k.autohome.com.cn/{series_id}/index_{page}.html?dimensionid={dimensionid}&order=0&yearid=0#listcontainer'


def get_snapshot(cwd: Path, url: str, session: str, retries: int = 2):
    last_err = ''
    for _ in range(retries + 1):
        r1 = run(f"agent-browser --session {session} open '{url}' >/dev/null", cwd)
        if r1.returncode == 0:
            r2 = run(f"agent-browser --session {session} snapshot -c", cwd)
            if r2.returncode == 0:
                out = r2.stdout
                if '查看完整口碑' in out and 'detail/view_' in out:
                    return out.splitlines()
            last_err = r2.stderr or r2.stdout
        else:
            last_err = r1.stderr or r1.stdout
        time.sleep(1)
    raise RuntimeError(last_err)


def get_page_html(cwd: Path, url: str, session: str, retries: int = 2):
    last_err = ''
    for _ in range(retries + 1):
        r1 = run(f"agent-browser --session {session} open '{url}' >/dev/null", cwd)
        if r1.returncode != 0:
            last_err = r1.stderr or r1.stdout
            time.sleep(1)
            continue
        r2 = run(f"agent-browser --session {session} eval 'document.documentElement.outerHTML'", cwd)
        if r2.returncode == 0 and r2.stdout.strip():
            return r2.stdout
        last_err = r2.stderr or r2.stdout
        time.sleep(1)
    raise RuntimeError(last_err)


def get_snapshot_any(cwd: Path, url: str, session: str, retries: int = 2):
    last_err = ''
    for _ in range(retries + 1):
        r1 = run(f"agent-browser --session {session} open '{url}' >/dev/null", cwd)
        if r1.returncode == 0:
            r2 = run(f"agent-browser --session {session} snapshot -c", cwd)
            if r2.returncode == 0 and r2.stdout.strip():
                return r2.stdout.splitlines()
            last_err = r2.stderr or r2.stdout
        else:
            last_err = r1.stderr or r1.stdout
        time.sleep(1)
    raise RuntimeError(last_err)


def detect_max_page(cwd: Path, series_id: int, dimensionid: int) -> int:
    html = get_page_html(cwd, url_for(series_id, dimensionid, 1), f'koubei_detect_{series_id}_{dimensionid}')
    html = unescape(html)

    candidates = set()

    for m in re.finditer(rf'/{series_id}/index_(\d+)\.html\?dimensionid={dimensionid}', html):
        candidates.add(int(m.group(1)))

    for m in re.finditer(r'ace-pagination__link[^>]*>(\d+)<', html):
        candidates.add(int(m.group(1)))

    for m in re.finditer(r'分页[^\n]{0,200}?共\s*(\d+)\s*页', html):
        candidates.add(int(m.group(1)))

    for m in re.finditer(r'共\s*(\d+)\s*页', html):
        candidates.add(int(m.group(1)))

    for m in re.finditer(r'尾页[^\n]{0,120}?index_(\d+)\.html', html):
        candidates.add(int(m.group(1)))

    candidates = {x for x in candidates if x >= 1}
    return max(candidates) if candidates else 1


def norm_user(raw: str) -> str:
    raw = raw.strip()
    if raw.endswith(' 风云X3L认证'):
        return raw[:-8].strip() + '_风云X3L认证'
    if raw.endswith(' 风云X3L'):
        return raw[:-6].strip()
    return raw


def extract_cards(lines):
    cards = []
    current = []
    in_target = False
    for line in lines:
        s = line.strip()
        if 'heading ' in s and '/detail/view_' not in s:
            pass
        if 'heading ' in s and current and any('查看完整口碑' in x for x in current):
            cards.append(current)
            current = []
        if current or ('heading ' in s and '[level=1]' in s):
            current.append(line)
            if '相关车系口碑推荐' in s:
                if current and any('查看完整口碑' in x for x in current):
                    cards.append(current)
                current = []
                break
    if current and any('查看完整口碑' in x for x in current):
        cards.append(current)

    detail_cards = []
    for c in cards:
        if any('https://k.autohome.com.cn/detail/view_' in x for x in c):
            detail_cards.append(c)

    # 去重：同一链接保留首次出现
    seen = set()
    uniq = []
    for c in detail_cards:
        link = ''
        for line in c:
            m = re.search(r'https://k\.autohome\.com\.cn/detail/view_[^\s]+', line)
            if m:
                link = m.group(0)
                break
        if link and link not in seen:
            seen.add(link)
            uniq.append(c)
    return uniq


def parse_card(lines, dim_name: str, page: int, cwd=None):
    row = {
        '口碑维度': dim_name,
        '抓取页码': page,
        '数据类型': '车主购车口碑',
        '用户名': '', '发表日期': '', '口碑标题': '', '综合口碑': '', '车型': '',
        '行驶里程': '', '电耗': '', '裸车购买价': '', '参考价格': '', '购买时间': '',
        '探店时间': '', '购买地点': '', '探店地点': '', '评价详情': '', '来源链接': ''
    }

    compact = ' '.join(l.strip() for l in lines)

    for i, l in enumerate(lines):
        s = l.strip()

        if s.startswith('- /url: https://i.autohome.com.cn/') and i + 1 < len(lines):
            m = re.match(r'- link "([^"]+)"', lines[i + 1].strip())
            if m and not row['用户名']:
                row['用户名'] = norm_user(m.group(1))

        if '发表口碑' in s:
            m = re.search(r'(20\d{2}-\d{2}-\d{2})\s+发表口碑', s)
            if m:
                row['发表日期'] = m.group(1)

        if '综合口碑评分' in s:
            m = re.search(r'综合口碑评分\s+([0-9.]+)', s)
            if m:
                row['综合口碑'] = m.group(1)

        m = re.match(r'- link "([^"]+)" \[ref=.*\]:$', s)
        if m and i + 1 < len(lines):
            nxt = lines[i + 1].strip()
            if 'https://k.autohome.com.cn/detail/view_' in nxt and not row['口碑标题']:
                row['口碑标题'] = m.group(1)

        m = re.match(r'- link "(20\d{2}款 [^"]+)"', s)
        if m and not row['车型']:
            row['车型'] = m.group(1)

        m = re.match(r'- listitem: (.+)', s)
        if m:
            val = m.group(1).strip()
            if val.endswith('行驶里程'):
                row['行驶里程'] = val[:-4].strip()
            elif '电耗' in val:
                row['电耗'] = val
            elif val.endswith('裸车购买价'):
                row['裸车购买价'] = val[:-5].strip()
            elif val.endswith('参考价格'):
                row['参考价格'] = val[:-4].strip()
                row['数据类型'] = '试驾探店口碑'
            elif val.endswith('购买时间'):
                row['购买时间'] = val[:-4].strip()
            elif val.endswith('探店时间'):
                row['探店时间'] = val[:-4].strip()
                row['数据类型'] = '试驾探店口碑'
            elif val.endswith('购买地点'):
                row['购买地点'] = val[:-4].strip()
            elif val.endswith('探店地点'):
                row['探店地点'] = val[:-4].strip()
                row['数据类型'] = '试驾探店口碑'

        for prefix, dtype in [
            ('- text: 满意 ', '车主购车口碑'),
            ('- text: 不满意 ', '车主购车口碑'),
            ('- text: 好评 ', '试驾探店口碑'),
            ('- text: 槽点 ', '试驾探店口碑')
        ]:
            if s.startswith(prefix):
                row['评价详情'] = s[len(prefix):].strip()
                row['数据类型'] = dtype

        if s in ('- text: 满意', '- text: 不满意') and not row['评价详情']:
            m = re.search(r'- text: (?:满意|不满意)\s+(.*?)\s+- listitem:', compact)
            if m:
                candidate = m.group(1).strip()
                if candidate and not candidate.startswith('- listitem:'):
                    row['评价详情'] = candidate
                    row['数据类型'] = '车主购车口碑'

        if s in ('- text: 好评', '- text: 槽点') and not row['评价详情']:
            m = re.search(r'- text: (?:好评|槽点)\s+(.*?)\s+- listitem:', compact)
            if m:
                candidate = m.group(1).strip()
                if candidate and not candidate.startswith('- listitem:'):
                    row['评价详情'] = candidate
                    row['数据类型'] = '试驾探店口碑'

        m = re.search(r'https://k\.autohome\.com\.cn/detail/view_[^\s]+', s)
        if m and not row['来源链接']:
            row['来源链接'] = m.group(0)

    if any([row['参考价格'], row['探店时间'], row['探店地点']]):
        row['数据类型'] = '试驾探店口碑'

    if not row['评价详情'] and row['来源链接'] and cwd is not None:
        try:
            detail_session = f"kd_{hashlib.md5(row['来源链接'].encode('utf-8')).hexdigest()[:12]}"
            detail_lines = get_snapshot_any(cwd, row['来源链接'], detail_session)
            detail_text = extract_detail_text(detail_lines, row['口碑维度'])
            if detail_text:
                row['评价详情'] = detail_text
        except Exception:
            pass

    valid = all(row[k] for k in ['用户名', '综合口碑', '车型', '来源链接', '评价详情'])
    return row, valid


def extract_detail_text(lines, dim_name: str) -> str:
    target_heading = dim_name
    collecting = False
    parts = []
    for raw in lines:
        s = raw.strip()
        if s.startswith(f'- heading "{target_heading}"'):
            collecting = True
            continue
        if collecting and s.startswith('- heading "'):
            break
        if collecting and s.startswith('- paragraph:'):
            text = s[len('- paragraph:'):].strip()
            text = text.strip('"').strip()
            if text and '上述内容的版权归' not in text:
                parts.append(text)
    return ' '.join(parts).strip()


def collect_dimension(
    cwd: Path,
    series_id: int,
    dimensionid: int,
    start_page: int,
    end_page: int,
    *,
    reporter=None,
    stage_key=None,
    stage_label=None,
    stage_weight=0,
    show_progress=False,
):
    dim_name = '最满意' if dimensionid == 10 else '最不满意'
    rows, bad = [], []
    page_link_counts = {}
    total_pages = end_page - start_page + 1
    if reporter and stage_key:
        reporter.start_stage(stage_key, stage_label or dim_name, weight=stage_weight, total=total_pages, message=f"开始抓取{stage_label or dim_name}")
    for idx, page in enumerate(progress_iter(range(start_page, end_page + 1), total=total_pages, desc=f'抓取{dim_name}', enabled=show_progress), start=1):
        try:
            lines = get_snapshot(cwd, url_for(series_id, dimensionid, page), f'koubei_{series_id}_{dimensionid}_{page}')
        except RuntimeError as e:
            bad.append({'口碑维度': dim_name, '抓取页码': page, '错误': str(e)})
            if reporter and stage_key:
                reporter.advance(idx, message=f"{stage_label or dim_name} {idx}/{total_pages}")
            continue

        cards = extract_cards(lines)
        page_link_counts[page] = len(cards)
        if not cards:
            bad.append({'口碑维度': dim_name, '抓取页码': page, '错误': '未解析到口碑卡片'})
            if reporter and stage_key:
                reporter.advance(idx, message=f"{stage_label or dim_name} {idx}/{total_pages}")
            continue

        for c in cards:
            row, valid = parse_card(c, dim_name, page, cwd=cwd)
            if valid:
                rows.append(row)
            else:
                bad.append(row)
        if reporter and stage_key:
            reporter.advance(idx, message=f"{stage_label or dim_name} {idx}/{total_pages}")
    if reporter and stage_key:
        reporter.finish_stage(f"{stage_label or dim_name}完成")
    return rows, bad, page_link_counts


def merge_aligned(sat_rows, unsat_rows):
    sat_map = {r['来源链接']: r for r in sat_rows}
    unsat_map = {r['来源链接']: r for r in unsat_rows}

    sat_links = set(sat_map)
    unsat_links = set(unsat_map)
    common = sorted(sat_links & unsat_links)
    only_sat = sorted(sat_links - unsat_links)
    only_unsat = sorted(unsat_links - sat_links)

    merged = []
    anomalies = []

    for link in common:
        a = sat_map[link]
        b = unsat_map[link]
        dtype = '试驾探店口碑' if any([
            a['数据类型'] == '试驾探店口碑',
            b['数据类型'] == '试驾探店口碑',
            a['参考价格'], b['参考价格'],
            a['探店时间'], b['探店时间'],
            a['探店地点'], b['探店地点'],
        ]) else '车主购车口碑'

        row = {
            '数据类型': dtype,
            '用户名': a['用户名'] or b['用户名'],
            '发表日期': a['发表日期'] or b['发表日期'],
            '口碑标题': a['口碑标题'] or b['口碑标题'],
            '综合口碑': a['综合口碑'] or b['综合口碑'],
            '车型': a['车型'] or b['车型'],
            '行驶里程': a['行驶里程'] or b['行驶里程'],
            '电耗': a['电耗'] or b['电耗'],
            '裸车购买价': a['裸车购买价'] or b['裸车购买价'],
            '参考价格': a['参考价格'] or b['参考价格'],
            '购买时间': a['购买时间'] or b['购买时间'],
            '探店时间': a['探店时间'] or b['探店时间'],
            '购买地点': a['购买地点'] or b['购买地点'],
            '探店地点': a['探店地点'] or b['探店地点'],
            '最满意': a['评价详情'],
            '最不满意': b['评价详情'],
            '来源链接': link,
            '最满意抓取页码': a['抓取页码'],
            '最不满意抓取页码': b['抓取页码'],
        }

        if not row['最满意'] or not row['最不满意']:
            anomalies.append({'type': 'missing_dimension_text', 'link': link, 'sat': a, 'unsat': b})
            continue

        if dtype == '车主购车口碑':
            row['参考价格'] = ''
            row['探店时间'] = ''
            row['探店地点'] = ''

        merged.append(row)

    for link in only_sat:
        anomalies.append({'type': 'missing_unsat', 'link': link, 'row': sat_map[link]})
    for link in only_unsat:
        anomalies.append({'type': 'missing_sat', 'link': link, 'row': unsat_map[link]})

    groups = {
        '购车口碑': [r for r in merged if r['数据类型'] == '车主购车口碑'],
        '试驾口碑': [r for r in merged if r['数据类型'] == '试驾探店口碑'],
    }
    return groups, anomalies, {'common': len(common), 'only_sat': only_sat, 'only_unsat': only_unsat}


def write_xlsx(out_path: Path, groups: dict):
    wb = Workbook()
    sheet_names = ['购车口碑', '试驾口碑']
    first = True
    for name in sheet_names:
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        ws.append(HEADERS)
        for r in groups.get(name, []):
            ws.append([r.get(h, '') for h in HEADERS])
        for c in ws[1]:
            c.font = Font(bold=True)
        for col in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','R','S']:
            ws.column_dimensions[col].width = 16
        ws.column_dimensions['O'].width = 90
        ws.column_dimensions['P'].width = 90
        ws.column_dimensions['Q'].width = 60
        for row in ws.iter_rows(min_row=2):
            row[14].alignment = Alignment(wrap_text=True, vertical='top')
            row[15].alignment = Alignment(wrap_text=True, vertical='top')
            row[16].alignment = Alignment(wrap_text=True, vertical='top')
    wb.save(out_path)


def make_reporter(args, *, model_name):
    progress_file = args.progress_file
    if progress_file and not Path(progress_file).is_absolute():
        progress_file = str(Path(progress_file).resolve())
    label = f"koubei-collector:{args.series_id}:{model_name}"
    return ProgressReporter(
        label=label,
        model_name=model_name,
        mode="autohome_collect",
        progress_file=progress_file,
        progress_webhook=args.progress_webhook,
        feishu_webhook=args.feishu_webhook,
        feishu_secret=args.feishu_secret,
    )


def write_validation_report(path: Path, report: dict):
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')


def main():
    parser = argparse.ArgumentParser(description='Export aligned Autohome koubei to xlsx')
    parser.add_argument('--series-id', type=int, required=True)
    parser.add_argument('--start-page', type=int, default=1)
    parser.add_argument('--end-page', type=int)
    parser.add_argument('--auto-detect-pages', action='store_true', help='Auto detect max page when end page is omitted')
    parser.add_argument('--output', required=True)
    parser.add_argument('--workdir', default='.')
    parser.add_argument('--strict-validate', action='store_true')
    parser.add_argument('--progress', action='store_true', help='Show terminal progress bar')
    parser.add_argument('--progress-file', help='Write progress state to JSON file for polling')
    parser.add_argument('--progress-webhook', help='POST progress state to a generic webhook')
    parser.add_argument('--feishu-webhook', help='POST progress updates to a Feishu incoming webhook')
    parser.add_argument('--feishu-secret', help='Feishu bot secret for signed webhooks')
    args = parser.parse_args()

    cwd = Path(args.workdir).resolve()
    show_progress = args.progress or sys.stderr.isatty()
    use_progress_sink = bool(args.progress_file or args.progress_webhook or args.feishu_webhook)
    reporter = make_reporter(args, model_name=f"series-{args.series_id}") if use_progress_sink else None

    if args.end_page is None:
        if reporter:
            reporter.start_stage('detect_pages', '探测总页数', weight=STAGE_WEIGHTS['detect_pages'], total=2, message='开始自动探测页数')
        if args.auto_detect_pages or args.start_page == 1:
            sat_max = detect_max_page(cwd, args.series_id, 10)
            unsat_max = detect_max_page(cwd, args.series_id, 11)
            args.end_page = max(sat_max, unsat_max)
            print(f'Auto-detected end page: {args.end_page} (sat={sat_max}, unsat={unsat_max})')
            if reporter:
                reporter.advance(2, message=f'已探测到总页数 {args.end_page}')
        else:
            raise SystemExit('--end-page is required when start-page is not 1; pass --auto-detect-pages if you want automatic detection')
        if reporter:
            reporter.finish_stage('总页数已探测')

    if args.start_page < 1 or args.end_page < args.start_page:
        raise SystemExit('invalid page range: require start-page >= 1 and end-page >= start-page')

    sat_rows, sat_bad, sat_pages = collect_dimension(
        cwd,
        args.series_id,
        10,
        args.start_page,
        args.end_page,
        reporter=reporter,
        stage_key='collect_sat',
        stage_label='采集最满意',
        stage_weight=STAGE_WEIGHTS['collect_sat'],
        show_progress=show_progress,
    )
    unsat_rows, unsat_bad, unsat_pages = collect_dimension(
        cwd,
        args.series_id,
        11,
        args.start_page,
        args.end_page,
        reporter=reporter,
        stage_key='collect_unsat',
        stage_label='采集最不满意',
        stage_weight=STAGE_WEIGHTS['collect_unsat'],
        show_progress=show_progress,
    )

    groups, anomalies, meta = merge_aligned(sat_rows, unsat_rows)

    out_path = Path(args.output).resolve()
    if reporter:
        reporter.start_stage('write_excel', '写出 Excel', weight=STAGE_WEIGHTS['write_excel'], total=1, message='开始写出 Excel')
    write_xlsx(out_path, groups)
    if reporter:
        reporter.finish_stage('Excel 已写出')

    report = {
        'ok': len(anomalies) == 0,
        'sat_total_raw': len(sat_rows),
        'unsat_total_raw': len(unsat_rows),
        'aligned_total': sum(len(v) for v in groups.values()),
        'groups': {k: len(v) for k, v in groups.items()},
        'page_link_counts': {
            '最满意': sat_pages,
            '最不满意': unsat_pages,
        },
        'raw_parse_anomalies': len(sat_bad) + len(unsat_bad),
        'alignment_meta': meta,
        'anomalies': anomalies,
    }
    report_path = out_path.with_suffix('.validation.json')
    if reporter:
        reporter.start_stage('write_validation', '写出校验报告', weight=STAGE_WEIGHTS['write_validation'], total=1, message='开始写出校验报告')
    write_validation_report(report_path, report)
    if reporter:
        reporter.finish_stage('校验报告已写出')
        reporter.finalize(message='导出完成')

    print(f'Wrote: {out_path}')
    print(f'Validation: {report_path}')
    print('Counts:')
    for k, v in groups.items():
        print(f'- {k}: {len(v)}')
    print(f'- raw sat: {len(sat_rows)}')
    print(f'- raw unsat: {len(unsat_rows)}')
    print(f'- anomalies: {len(anomalies)}')
    if report['ok']:
        print('- validation: OK')
    else:
        print('- validation: FAILED')
        if args.strict_validate:
            raise SystemExit(2)


if __name__ == '__main__':
    main()
